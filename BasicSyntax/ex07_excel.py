# coding=utf-8
import numpy as np
import pandas as pd
import openpyxl as xl
import os
import statsmodels.api as sm
import statsmodels.tsa.stattools as stools



data_file_e = r'D:\code\Projects\Test\Lectures\XTL\Energy.xlsx'

data_path_e =data_file_e
# 直接保存 dataframe
data = pd.DataFrame(data=[[1,2,3],[4,5,6]])



# 读取 excel
df = pd.read_excel(data_path_e)
print(df)
x = df.iloc[:,2:3]
y = df.iloc[:,1:2]
print(x)
print(y)
X = sm.add_constant(x)
model = sm.OLS(y,X)
result_ols = model.fit()
print(result_ols.summary())
#print(df.iloc[:,3:5])   # 取其中指定列
#print(df.loc[:,'陕西'])

r = stools.adfuller(y)
print(r)

# 操作 Excel
# 新建 workbook
book_b = xl.Workbook()
book_b.create_sheet('Lecture')

book_b.close()

# 读取 excel表格

book_e = xl.load_workbook(data_path_e)
sheet = book_e.worksheets[0]
values = sheet['A1':'C28']
df = pd.DataFrame(values)
#print(df)

book_e.close()











